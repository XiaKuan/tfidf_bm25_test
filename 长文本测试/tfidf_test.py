#!/usr/bin/env python
#-*- coding:utf-8 -*-
import jieba
import jieba.analyse
import jieba.posseg as pseg

import nltk
import math
import string
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
# 根据列的数字返回字母
get_column_letter(2) # B
# 根据字母返回列的数字
column_index_from_string('D') # 4

import json
# from　nltk.corpus import stopwords
from collections import Counter




def get_tokens(text):
    # lower = text.lower()
    # remove_punctuation_map = dict((ord(char), None) for char in string.punctuation)
    # no_punctuation = lower.translate(remove_punctuation_map)
    tokens = jieba.lcut(text,cut_all=True)

    return tokens


def tf(word, count):
    return count[word] / sum(count.values())

def n_containing(word, count_list):
    return sum(1 for count in count_list if word in count)

def idf(word, count_list):
    return math.log(len(count_list)) / (1 + n_containing(word, count_list))

def tfidf(word, count, count_list):
    return tf(word, count) * idf(word, count_list)

def avgDl(text,texts):
    # 返回文档长度和平均长度的比值
    avgL = sum(len(i) for i in texts)/len(texts)
    return abs(len(text))/avgL

def bm25tf(word, count,l,b=0.75,k=1.2):
    """ 使用bm25算法得到的tf """
    return ((k+1)*tf(word, count))/(k*(1.0-b+b*l)+tf(word, count))

def bm25tfidf(word, count, count_list,l,b=0.75,k=1.2):
    """ 使用bm25算法得到的tf-idf 
    k：这个参数控制着词频结果在词频饱和度中的上升速度。默认值为1.2。
    值越小饱和度变化越快，值越大饱和度变化越慢。
    b：这个参数控制着字段长归一值所起的作用，
    0.0会禁用归一化，1.0会启用完全归一化。默认值为0.75
    """
    
    return bm25tf(word, count,l,b,k) * idf(word, count_list)

def count_term(text):
    """ 
    返回一个counter类,形如
    Counter({'你好': 2, '中国': 2, '再见': 2})
     """
    tokens = get_tokens(text)
    # filtered = [w for w in tokens if not w in stopwords.words('english')]
    # stemmer = PorterStemmer()
    # stemmed = stem_tokens(filtered, stemmer)
    # 词干抽取，删除
    count = Counter(tokens)
    return count

def load_xlsx(filename):
    """ 
    读取xlsx数据，返回可用的文本数据 
    """
    wb = openpyxl.load_workbook(filename)
    sheet_name = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheet_name[0])
    return [sheet['A'+str(row)].value for row in range(2,sheet.max_row)]

def load_json(filename):
    """ 
    加载json格式,处理成为可用的文本语料数组
    返回值包含所有文本语料的数组和一个对应的编号数组
    """
    with open('sort_cluster_k_300.json','r',encoding='utf-8') as fp:
        json_data = json.load(fp)
    # 导入测试语料
    texts = []
    keyindex = []
    key_num = []
    for data in json_data:
        text = ''
        for key,values in data.items():
            if key != 'num':            
                for i in values:
                    text = text+i
                keyindex.append(key)
            else:
                key_num.append(values)
                
        texts.append(text)
    return texts,keyindex,key_num

def main():
    filename = 'data.xlsx'
    texts = load_xlsx(filename)        
    # 读取文件
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet'
    wb.get_sheet_by_name('Sheet')
    sheet['A1'] = '编号'
    sheet['B1'] = '聚类数量'
    # 和编写结果文件抬头


    countlist = []
    L = []
    for text in texts[0:10000]:
        countlist.append(count_term(text))
        L.append(avgDl(text,texts))
        print('{:.2f}%'.format(len(L)/len(texts)*100))

    for i, count in enumerate(countlist):
        # print("Top words in document {}".format(i + 1))
        l = L[i]
        # 取得对应的文档长度与平均文档长度的比值
        # scores = {word: tfidf(word, count, countlist) for word in count}
        scores = {word: bm25tfidf(word, count, countlist,l,b=0.75,k=1.2) for word in count}
        sorted_words = sorted(scores.items(), key = lambda x: x[1], reverse=True)
        tfcount = [bm25tf(sorted_words[j][0],count,l,b=0.75,k=1.2) for j in range(len(sorted_words))]
        idfcount = [idf(sorted_words[j][0],countlist) for j in range(len(sorted_words))]

        # print(sorted_words)
        # for word, score in sorted_words[:5]:
        #     print("\tWord: {}, TF-IDF: {}".format(word, round(score, 5)))
        # sheet['A'+str(i+2)]=keyindex[i]
        # sheet['B'+str(i+2)]=key_num[i]
        print('{:.2f}%'.format(i/len(countlist)*100))
        sheet[get_column_letter(2)+str(i+2)]=texts[i]
        for j in range(0,len(sorted_words),4):
            sheet[get_column_letter(j+3)+str(i+2)]=sorted_words[j][0]
            sheet[get_column_letter(j+4)+str(i+2)]=sorted_words[j][1]
            sheet[get_column_letter(j+5)+str(i+2)]=tfcount[j]
            sheet[get_column_letter(j+6)+str(i+2)]=idfcount[j]


    wb.save("result1.xlsx")   

        

if __name__ == "__main__":
    main()